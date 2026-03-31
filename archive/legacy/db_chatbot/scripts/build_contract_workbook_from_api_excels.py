#!/usr/bin/env python3
"""Compose data/brand_data.xlsx from api_data Excel outputs."""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    headers = [str(v).strip() if v is not None else "" for v in header]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        rec = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            rec[key] = row[i] if i < len(row) else None
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            out.append(rec)
    return out


def to_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None


def parse_range_mid(value: Any) -> int | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    if not nums:
        return None
    vals = [float(n) for n in nums]
    return int(round(sum(vals) / len(vals)))


def write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def main() -> int:
    parser = argparse.ArgumentParser(description="Build contract workbook from api_data excels.")
    parser.add_argument(
        "--api-root",
        type=Path,
        default=Path("db_chatbot/api_data"),
        help="Root folder containing API output excels.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/brand_data.xlsx"),
        help="Output workbook path used by load_validate_data contract.",
    )
    args = parser.parse_args()

    api = args.api_root
    brand_list = read_xlsx_rows(api / "brand_list_info/output/brand_list_info.xlsx")
    frcs = read_xlsx_rows(api / "brand_frcs_stats/output/brand_frcs_stats.xlsx")
    fntn = read_xlsx_rows(api / "brand_fntn_stats/output/brand_fntn_stats.xlsx")
    interior = read_xlsx_rows(api / "brand_interior_cost/output/brand_interior_cost.xlsx")

    # brand_master
    master_by_id: dict[int, dict[str, Any]] = {}
    for r in brand_list:
        brand_id = to_int(r.get("brand_id"))
        if brand_id is None:
            continue
        rec = {
            "brand_id": brand_id,
            "brand_name": r.get("brandNm"),
            "corp_name": r.get("corpNm"),
            "category_l": r.get("indutyLclasNm"),
            "category_m": r.get("indutyMlsfcNm"),
            "start_date": r.get("jngBizStrtDate"),
        }
        prev = master_by_id.get(brand_id)
        if prev is None:
            master_by_id[brand_id] = rec
            continue
        prev_score = sum(1 for v in prev.values() if v is not None and str(v).strip() != "")
        rec_score = sum(1 for v in rec.values() if v is not None and str(v).strip() != "")
        if rec_score >= prev_score:
            master_by_id[brand_id] = rec
    brand_master_rows = list(master_by_id.values())

    # brand_year_stats
    stats_by_key: dict[tuple[int, int], dict[str, Any]] = {}
    by_brand_year: dict[int, dict[int, dict[str, Any]]] = defaultdict(dict)
    for r in frcs:
        brand_id = to_int(r.get("brand_id"))
        year = to_int(r.get("yr"))
        if brand_id is None or year is None:
            continue
        store_count = to_int(r.get("frcsCnt")) or 0
        new_count = to_int(r.get("newFrcsRgsCnt")) or 0
        closed_count = to_int(r.get("ctrtCncltnCnt")) or 0
        rec = {
            "brand_id": brand_id,
            "year": year,
            "store_count": store_count,
            "new_store_count": new_count,
            "closed_store_count": closed_count,
            "avg_sales_amt": to_int(r.get("avrgSlsAmt")),
            "net_store_change": new_count - closed_count,
            "store_growth_rate": None,
            "closure_rate": (closed_count / store_count) if store_count > 0 else 0.0,
            "churn_rate": ((new_count + closed_count) / store_count) if store_count > 0 else 0.0,
        }
        by_brand_year[brand_id][year] = rec

    for _, year_map in by_brand_year.items():
        years = sorted(year_map.keys())
        for i, y in enumerate(years):
            row = year_map[y]
            net = row["net_store_change"] or 0
            if i > 0:
                prev_store_count = float(year_map[years[i - 1]]["store_count"] or 0)
            else:
                prev_store_count = float((row["store_count"] or 0) - net)
            if prev_store_count > 0:
                row["store_growth_rate"] = net / prev_store_count
            else:
                row["store_growth_rate"] = 0.0
            stats_by_key[(row["brand_id"], row["year"])] = row

    brand_year_stats_rows = list(stats_by_key.values())

    # brand_store_types (derive one Standard type from interior area)
    type_by_brand: dict[int, dict[str, Any]] = {}
    for r in interior:
        brand_id = to_int(r.get("brand_id"))
        area = r.get("storCrtraAr")
        if brand_id is None:
            continue
        rec = {
            "brand_id": brand_id,
            "store_type": "Standard",
            "standard_area_pyeong": float(area) if area is not None else None,
        }
        type_by_brand[brand_id] = rec
    brand_store_types_rows = list(type_by_brand.values())

    # brand_store_type_costs
    cost_rows: list[dict[str, Any]] = []
    interior_mid_by_brand_year: dict[tuple[int, int], int] = {}
    for r in interior:
        brand_id = to_int(r.get("brand_id"))
        year = to_int(r.get("jngBizCrtraYr"))
        if brand_id is None or year is None:
            continue
        mid = parse_range_mid(r.get("intrrAmtScopeVal"))
        if mid is not None:
            interior_mid_by_brand_year[(brand_id, year)] = mid

    for r in fntn:
        brand_id = to_int(r.get("brand_id"))
        year = to_int(r.get("yr"))
        if brand_id is None or year is None:
            continue
        mappings = [
            ("initial_fee", to_int(r.get("jngBzmnJngAmt"))),
            ("education", to_int(r.get("jngBzmnEduAmt"))),
            ("other", to_int(r.get("jngBzmnEtcAmt"))),
            ("guarantee", to_int(r.get("jngBzmnAssrncAmt"))),
            ("total_initial_cost", to_int(r.get("smtnAmt"))),
        ]
        mid = interior_mid_by_brand_year.get((brand_id, year))
        if mid is not None:
            mappings.append(("interior", mid))
        for cost_category, amount in mappings:
            if amount is None:
                continue
            cost_rows.append(
                {
                    "brand_id": brand_id,
                    "year": year,
                    "store_type": "Standard",
                    "cost_category": cost_category,
                    "amount": amount,
                }
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "brand_master"
    write_sheet(
        ws,
        ["brand_id", "brand_name", "corp_name", "category_l", "category_m", "start_date"],
        brand_master_rows,
    )

    ws = wb.create_sheet("brand_year_stats")
    write_sheet(
        ws,
        [
            "brand_id",
            "year",
            "store_count",
            "new_store_count",
            "closed_store_count",
            "avg_sales_amt",
            "net_store_change",
            "store_growth_rate",
            "closure_rate",
            "churn_rate",
        ],
        brand_year_stats_rows,
    )

    ws = wb.create_sheet("brand_store_types")
    write_sheet(
        ws,
        ["brand_id", "store_type", "standard_area_pyeong"],
        brand_store_types_rows,
    )

    ws = wb.create_sheet("brand_store_type_costs")
    write_sheet(
        ws,
        ["brand_id", "year", "store_type", "cost_category", "amount"],
        cost_rows,
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"Wrote workbook: {args.output}")
    print(f"brand_master rows: {len(brand_master_rows)}")
    print(f"brand_year_stats rows: {len(brand_year_stats_rows)}")
    print(f"brand_store_types rows: {len(brand_store_types_rows)}")
    print(f"brand_store_type_costs rows: {len(cost_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
